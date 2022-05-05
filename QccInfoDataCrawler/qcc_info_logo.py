"""
# -*- coding: utf-8 -*-
#
# Copyright (C) 2021 #
# @Time    : 2022/4/29 15:23
# @Author  : zicliang
# @Email   : hybpjx@163.com
# @File    : qcc_info_logo.py
# @Software: PyCharm
"""
import os
import time
import requests
from openpyxl import load_workbook
from utils.Mixins import Mixins
from pyquery import PyQuery as pq
from settings.log_conf import *


class Request(object):
    def __init__(self):

        m = Mixins()
        fp = open("cookies/cookie.txt", "r", encoding="utf-8")
        cookies = fp.read()
        fp.close()
        self.cookies = m.cookies_dict(cookies)

        wb = load_workbook('source/qcc_info_logo.xlsx')
        # 得到sheet页的对象
        self.sheet = wb['whole']
        # 判断语句 如果不存在则创建
        if not os.path.exists("QccLogo"):
            os.mkdir("QccLogo")

        self.source_path = '../QccLogo'
        # 存放图片url地址
        self.url_list = []

        self.title_list = []

    def main(self):
        url_list = []
        for col in self.sheet.iter_cols(min_col=3, max_col=3, min_row=2):
            for cell in col:
                url_list.append(cell.value)

        name_list = []
        # min 是 从第几列开始 max 是从第几列结束
        for col in self.sheet.iter_cols(min_col=1, max_col=1, min_row=2):
            for cell in col:
                name_list.append(cell.value)

        for url, picName in zip(url_list, name_list):
            try:
                time.sleep(3)

                # cookie.txt 伪装
                headers = {
                    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
                    'accept-encoding': 'gzip, deflate, br',
                    'accept-language': 'zh-CN,zh;q=0.9',
                    'cache-control': 'max-age=0',
                    'sec-ch-ua': '"Google Chrome";v="96", "Chromium";v="96", ";Not A Brand";v="99"',
                    'sec-ch-ua-mobile': '?0',
                    'sec-ch-ua-platform': '"Windows"',
                    'sec-fetch-dest': 'document',
                    'sec-fetch-mode': 'navigate',
                    'sec-fetch-site': 'same-origin',
                    'sec-fetch-user': '?1',
                    'upgrade-insecure-requests': '1',
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36'
                }
                # 发送请求
                session = requests.session()
                response = session.get(url=url, headers=headers, cookies=self.cookies)
                # 检测状态码
                status = response.status_code
                if status == 200:
                    # 得到源码
                    page_source = response.text
                    # 生成一个pyquery
                    doc = pq(page_source)
                    # 拿到 logo 图片
                    img_url = doc(".img img").attr("src")
                    # 由解析可知 logo 图片带水印 去掉问号就是去除水印
                    img_url = str(img_url).split("?")[0]

                    img_name = doc(".title.copy-value").text()
                    # print(img_name,"\n",img_url)
                    # 对图片发起请求
                    img_data_response = session.get(url=img_url, headers=headers,cookies=self.cookies)
                    # 拿到二进制数据
                    img_data = img_data_response.content

                    if picName == '-':
                        img_path = self.source_path + '/' + img_name
                        with open(img_path, 'wb') as fp:
                            fp.write(img_data)
                    else:
                        img_path = self.source_path + '/' + picName
                        with open(img_path, 'wb') as fp:
                            fp.write(img_data)

                    logger.info('{} >>>>>>>>>>> is ok'.format(img_url))

                else:
                    logger.error(f"网址：{url} 状态码为：{response.status_code} 请更换 cookie.txt")
            except Exception as e:
                logger.critical(e)


if __name__ == '__main__':
    Request().main()
