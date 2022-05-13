"""
# -*- coding: utf-8 -*-
#
# Copyright (C) 2021 #
# @Time    : 2022/3/1 10:15
# @Author  : zicliang
# @Email   : hybpjx@163.com
# @File    : main.py
# @Software: PyCharm
"""
from email import header
import time
import pymongo
import requests
from lxml import etree
from openpyxl import load_workbook
from settings.log_conf import *
import pandas as pd

import urllib
# myclient = pymongo.MongoClient("mongodb://localhost:27017/")
# mydb = myclient['QccUrl']
# mycol = mydb['titleUrlNews1']

# wb = load_workbook('./source/qcc_info_url.xlsx')

# # 得到sheet页的对象
# sheet = wb['Sheet1']

# 存放图片url地址
url_list = []


# def Proxies() -> dict:
#     # 要访问的目标页面
#     targetUrl = "http://ip4.hahado.cn/ip"
#
#     # 代理服务器
#     proxyHost = "ip4.hahado.cn"
#     proxyPort = "42228"
#     # 191709 191709
#     # 代理隧道验证信息
#     proxyUser = "191715"
#     proxyPass = "191715"
#
#     proxyMeta = "http://%(user)s:%(pass)s@%(host)s:%(port)s" % {
#         "host": proxyHost,
#         "port": proxyPort,
#         "user": proxyUser,
#         "pass": proxyPass,
#     }
#
#     proxies = {
#         "http://": proxyMeta,
#         "https://": proxyMeta,
#     }
#
#     return proxies

session = requests.Session()

def fetch_session():
    session.get('https://www.qcc.com')


def request():
    df = pd.read_csv('./name_list.csv', header=0)
    name_list = df['企业名称']
    # min 是 从第几列开始 max 是从第几列结束
    # for col in sheet.iter_cols(min_col=1, max_col=1, min_row=2):
    #     for cell in col:
    #         name_list.append(cell.value)

    result = []
    count = 1
    for name in name_list:
        url = 'https://www.qcc.com/web/search?key=' + str(name)

        # cookie.txt 伪装
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'zh-CN,zh;q=0.9',
            'cache-control': 'no-cache',
            # 'cookie.txt': 'qcc_did=33213b33-97cf-47f6-bfb3-4ada13d561f0; UM_distinctid=17f43dbebb09a5-00ca89c96828dc-a3e3164-1fa400-17f43dbebb110dc; QCCSESSID=3bac48cc49954a47bcedb6f3e3; acw_tc=3ad8379e16469653004761431eddd7792cc1d7911d382e1c6f9eb8001a; CNZZDATA1254842228=787216255-1646101019-%7C1646965021',
            'pragma': 'no-cache',
            'referer': f'https://www.qcc.com/web/search?key={urllib.parse.quote(name, safe="")}',
            'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="99", "Google Chrome";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-user': '?1',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36',
            'img': ''
        }
        # 发送请求
        response = session.get(url, headers=headers)
        status = response.status_code
        if status == 200:
            # 得到源码
            time.sleep(1.75)
            page_source = response.text
            # 得到一个tree对象
            tree = etree.HTML(page_source)
            # print(page_source)
            # xpath解析 得到 titile地址
            try:
                title_url = tree.xpath('//a[@class="title copy-value"]/@href')[0]
                title_name = tree.xpath('//a[@class="title copy-value"]/span')[0]
                credit_code = tree.xpath('//div[@class="relate-info"]/div[@class="rline over-rline"]//span[@class="copy-value"]')[0]
                address = tree.xpath('//div[@class="relate-info"]/div[@class="rline"][not(contains (@class, "over-rline"))]//span[@class="copy-value"]')[0]
                title_name = ''.join(title_name.xpath('.//text()'))
                url_list.append(title_url)
                credit_code = credit_code.xpath('.//text()')[0]
                address = ''.join(address.xpath('.//text()'))

                mydict = {"name": title_name, "credit_code": credit_code, "url": title_url, "address": address}
                result.append(mydict)
                print(mydict)
                time.sleep(0.5)
                # # mycol.insert_one(mydict)
                # logger.info(f"企业：{name},地址为：{title_url},写入成功")
            except:
                logger.error(f'{name} 爬取失败 能正常访问页面 可能是xpath有所改动，请修改xpath后重试')

        else:
            logger.error(f'{name} 爬取失败 状态码为：{response.status_code} 请更换cookie 或是稍后 再重试')
            raise ConnectionError(f'{name} 状态码为：{response.status_code} 爬取失败 或是稍后 请更换cookie 再重试')
        if len(result) == 100:
            pd.DataFrame(result).to_csv(f'./result_{count}.csv')
            result = []
            count = count + 1


if __name__ == '__main__':
    fetch_session()
    request()

    # for i, url in enumerate(url_list):
    #     sheet.cell(row=i + 2, column=2).value = url
